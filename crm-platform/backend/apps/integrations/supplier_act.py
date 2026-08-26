# -*- coding: utf-8 -*-
"""Розбір накладної/рахунку постачальника (.xls BIFF / .xlsx).
Підтримує різні формати (Корженевський «Замовлення покупця», Єлдинов «Рахунок на оплату» тощо):
шапка таблиці зіставляється НЕЧІТКО (Товар* / Кіл* / Ці* / Сума), колонки — за позицією в шапці.
Закупівельна ціна за од. = Сума / Кількість (або колонка «Ціна»). Повертає рядки для звірки зі складом.
"""
import re

_UA_MONTHS = {
    "січня": 1, "лютого": 2, "березня": 3, "квітня": 4, "травня": 5, "червня": 6,
    "липня": 7, "серпня": 8, "вересня": 9, "жовтня": 10, "листопада": 11, "грудня": 12,
}


def _ua_date(line):
    m = re.search(r"(\d{1,2})\s+([а-яіїєґ']+)\s+(\d{4})", line, re.I)
    if m and m.group(2).lower() in _UA_MONTHS:
        return "%02d.%02d.%s" % (int(m.group(1)), _UA_MONTHS[m.group(2).lower()], m.group(3))
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", line)
    if m:
        return "%02d.%02d.%s" % (int(m.group(1)), int(m.group(2)), m.group(3))
    return None


def _f(v):
    try:
        return float(str(v).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def _rows_from_bytes(data):
    """Читає .xls (BIFF, xlrd) або .xlsx (openpyxl). Повертає список рядків (список клітинок-рядків)."""
    if data[:4] == b"\xd0\xcf\x11\xe0":  # OLE2 = старий .xls
        import xlrd
        wb = xlrd.open_workbook(file_contents=data)
        ws = wb.sheet_by_index(0)
        return [[str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)] for r in range(ws.nrows)]
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(["" if v is None else str(v).strip() for v in row])
    return out


# слова, що позначають кінець таблиці товарів
_END_MARKERS = ("Разом", "Всього до", "Всьго до", "Без податку", "ПДВ", "Всього наймен")


def parse_korzh(file_contents):
    """file_contents — bytes .xls/.xlsx. Повертає dict акта постачальника."""
    rows = _rows_from_bytes(file_contents)

    out = {"invoice_number": None, "invoice_date": None,
           "supplier": {"name": None, "iban": None, "ipn": None},
           "lines": [], "total": None}
    name_col = qty_col = price_col = sum_col = code_col = None
    header_r = None

    for r, row in enumerate(rows):
        line = " ".join(x for x in row if x)

        # ── номер накладної/рахунку (різні формулювання) ──
        if not out["invoice_number"]:
            m = re.search(r"(?:Рахунок на оплату|Замовлення покупця|Видаткова накладна|Накладна|Рахунок)\s*№\s*([^\s,]+)", line, re.I)
            if m:
                out["invoice_number"] = m.group(1)
                out["invoice_date"] = _ua_date(line)

        # ── постачальник (Постачальник / Одержувач) ──
        if not out["supplier"]["name"] and ("Постачальник" in line or "Одержувач" in line):
            m2 = re.search(r"(?:Постачальник|Одержувач)\s*:?\s*(.+)", line)
            if m2:
                nm = re.split(r",|ІПН|ЄДРПОУ|ДРФО|р/р|IBAN|UA\d", m2.group(1))[0].strip()
                if nm and len(nm) >= 4:
                    out["supplier"]["name"] = nm[:160]
        m = re.search(r"(UA\d{27})", line)
        if m and not out["supplier"]["iban"]:
            out["supplier"]["iban"] = m.group(1)
        m = re.search(r"(?:ІПН|ЄДРПОУ|ДРФО)\s*(\d{8,12})", line)
        if m and not out["supplier"]["ipn"]:
            out["supplier"]["ipn"] = m.group(1)

        # ── підсумок із фрази "на суму X грн" (найнадійніше) ──
        m = re.search(r"на суму\s+([\d\s.,]+)\s*грн", line, re.I)
        if m and out["total"] is None:
            out["total"] = _f(m.group(1))

        # ── шапка таблиці (нечітке зіставлення) ──
        if header_r is None:
            has_name = any(str(v).strip().startswith("Товар") for v in row)
            has_qty = any("Кіл" in str(v) for v in row)
            if has_name and has_qty:
                for c, val in enumerate(row):
                    vs = str(val).strip()
                    if name_col is None and vs.startswith("Товар"):
                        name_col = c
                    elif qty_col is None and "Кіл" in vs:
                        qty_col = c
                    elif vs == "Код":
                        code_col = c
                    elif price_col is None and vs.startswith("Ці"):
                        price_col = c
                    elif "Сума" in vs:
                        sum_col = c  # остання «Сума»
                header_r = r
                continue

        # ── рядки товарів ──
        if header_r is not None and name_col is not None and r > header_r:
            if any(str(x).startswith(_END_MARKERS) for x in row):
                if out["total"] is None:
                    nums = [_f(x) for x in row if _f(x) is not None]
                    if nums:
                        out["total"] = nums[-1]
                break
            name = row[name_col] if name_col < len(row) else ""
            qty = _f(row[qty_col]) if (qty_col is not None and qty_col < len(row)) else None
            if name and qty is not None:
                sm = _f(row[sum_col]) if (sum_col is not None and sum_col < len(row)) else None
                pr = _f(row[price_col]) if (price_col is not None and price_col < len(row)) else None
                if pr is None and sm is not None and qty:
                    pr = round(sm / qty, 4)
                if sm is None and pr is not None:
                    sm = round(pr * qty, 2)
                out["lines"].append({"name": name, "qty": qty, "price": pr, "sum": sm})

    if out["total"] is None and out["lines"]:
        out["total"] = round(sum(l["sum"] or 0 for l in out["lines"]), 2)
    return out


if __name__ == "__main__":
    import sys
    import json
    data = open(sys.argv[1], "rb").read()
    print(json.dumps(parse_korzh(data), ensure_ascii=False, indent=1))


def parse_generic_invoice(raw, name=""):
    """Універсальний парсер рахунку-фактури постачальника (.xls/.xlsx) — коли специфічний
    парсер конкретного постачальника не спрацював. Шукає рядок-заголовок таблиці
    (Товар/Кількість/Ціна/Сума), знімає позиції до підсумку. Виправляє cp1251 у старих .xls."""
    import io as _io, re as _re
    rows = []
    low = (name or "").lower()
    try:
        if low.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
            ws = wb[wb.sheetnames[0]]
            for r in ws.iter_rows(values_only=True):
                rows.append(list(r))
        else:
            import xlrd
            wb = xlrd.open_workbook(file_contents=raw, encoding_override="cp1251")
            sh = wb.sheet_by_index(0)
            for ri in range(sh.nrows):
                rows.append([sh.cell_value(ri, ci) for ci in range(sh.ncols)])
    except Exception:
        return {"lines": [], "invoice_number": None, "invoice_date": None, "total": None, "supplier": {}}

    def _s(x):
        return str(x).strip() if x is not None else ""

    def _num(x):
        t = _s(x).replace("\xa0", "").replace(" ", "").replace(",", ".")
        t = _re.sub(r"[^0-9.\-]", "", t)
        try:
            return float(t) if t not in ("", "-", ".") else None
        except Exception:
            return None

    hdr = None
    col = {}
    for i, r in enumerate(rows):
        j = " ".join(_s(c).lower() for c in r)
        if ("товар" in j or "наймен" in j or "номенклат" in j) and any(_qk in j for _qk in ("кільк", "к-сть", "к-ть", "кіл-", "кіл.", "колич", "кол-во", "кол-сть")) and ("сума" in j or "сумма" in j):
            for ci, c in enumerate(r):
                t = _s(c).lower()
                if ("товар" in t or "наймен" in t or "номенклат" in t) and "name" not in col:
                    col["name"] = ci
                elif any(_qk in t for _qk in ("кільк", "к-сть", "к-ть", "кіл-", "кіл.", "колич", "кол-во", "кол-сть")):
                    col["qty"] = ci
                elif ("ціна" in t or "цена" in t) and "price" not in col:
                    col["price"] = ci
                elif ("сума" in t or "сумма" in t) and "sum" not in col:
                    col["sum"] = ci
            hdr = i
            break
    if hdr is None or "name" not in col:
        return {"lines": [], "invoice_number": None, "invoice_date": None, "total": None, "supplier": {}}

    lines = []
    total = None
    for r in rows[hdr + 1:]:
        nm = _s(r[col["name"]]) if col["name"] < len(r) else ""
        j = " ".join(_s(c).lower() for c in r)
        if any(k in j for k in ("разом", "всього", "итого", "загалом")):
            for c in r:
                v = _num(c)
                if v and (total is None or v > total):
                    total = v
            continue
        if not nm:
            continue
        qty = _num(r[col["qty"]]) if col.get("qty") is not None and col["qty"] < len(r) else None
        price = _num(r[col["price"]]) if col.get("price") is not None and col["price"] < len(r) else None
        ssum = _num(r[col["sum"]]) if col.get("sum") is not None and col["sum"] < len(r) else None
        if qty is None and ssum is None:
            continue
        lines.append({"name": nm[:200], "qty": qty or 0, "price": price or 0, "sum": ssum})

    alltext = "\n".join(" ".join(_s(c) for c in r) for r in rows)
    mi = _re.search(r"[Рр]ахунок[-\s]*фактура\s*[№N]?\s*([^\s,;]+)", alltext)
    inv = mi.group(1).strip() if mi else None
    if not inv:   # «Рахунок на оплату № 555» тощо
        mi2 = _re.search(r"[Рр]ахунок[^№\n]{0,40}№\s*([0-9][^\s,;]*)", alltext)
        inv = mi2.group(1).strip() if mi2 else None
    md = _re.search(r"\d{2}\.\d{2}\.\d{4}", alltext)
    dt = md.group(0) if md else None
    if not dt:   # «від 26 серпня 2026 р.»
        md2 = _re.search(r"від\s+(\d{1,2}\s+[а-яіїєґ']+\s+\d{4})", alltext, _re.I)
        if md2:
            dt = _ua_date(md2.group(1))
    if total is None and lines:
        total = round(sum((l["sum"] or 0) for l in lines), 2)
    # реквізити ПОСТАЧАЛЬНИКА (продавця): від рядка "Постачальник" до "Одержувач/Покупець"
    _sup = {"name": None, "iban": None, "edrpou": None, "ipn": None}
    _ps = _pe = None
    for _i, _r in enumerate(rows):
        _jl = " ".join(_s(c) for c in _r).lower()
        if _ps is None and "постачальник" in _jl:
            _ps = _i
        elif _ps is not None and _pe is None and ("одержувач" in _jl or "покупець" in _jl or "платник" in _jl):
            _pe = _i
    _blk = "\n".join(" ".join(_s(c) for c in _r) for _r in rows[(_ps or 0):(_pe if _pe is not None else (hdr if hdr is not None else len(rows)))])
    _mi = _re.search(r"UA\d{27}", _blk.replace(" ", ""))
    if _mi:
        _sup["iban"] = _mi.group(0)
    _me = _re.search(r"ЄДРПОУ[:\s]*([0-9]{6,10})", _blk)
    if _me:
        _sup["edrpou"] = _me.group(1)
    _mp = _re.search(r"[ІИ]ПН[:\s]*([0-9]{8,12})", _blk)
    if _mp:
        _sup["ipn"] = _mp.group(1)
    # fallback по всьому тексту (коли блок постачальника не виділився через слово «Постачальника» у шапці-інструкції)
    if not _sup["ipn"]:
        _m2 = _re.search(r"[ІИ]ПН\s*([0-9]{10,12})", alltext)
        if _m2:
            _sup["ipn"] = _m2.group(1)
    if not _sup["edrpou"]:
        _m2 = _re.search(r"ЄДРПОУ\s*([0-9]{6,10})", alltext)
        if _m2:
            _sup["edrpou"] = _m2.group(1)
    if not _sup["iban"]:
        _m2 = _re.search(r"UA\d{27}", alltext.replace(" ", ""))
        if _m2:
            _sup["iban"] = _m2.group(0)
    # назва — з рядка «Постачальник: <назва>, ІПН …»
    _mn = _re.search(r"[Пп]остачальник\s*:\s*(.+)", alltext)
    if _mn:
        _nm = _re.split(r",\s*(?:ІПН|ИНН|код|ЄДРПОУ|р/р|IBAN)", _mn.group(1).strip())[0].strip().strip('"').strip()
        if len(_nm) > 3:
            _sup["name"] = _nm[:160]
    if not _sup["name"] and _ps is not None:
        for _r in rows[_ps:_ps + 2]:
            for _c in _r:
                _t = _s(_c)
                if _t and "постачальник" not in _t.lower() and len(_t) > 3:
                    _sup["name"] = _t[:160]
                    break
            if _sup["name"]:
                break
    return {"lines": lines, "invoice_number": inv, "invoice_date": dt, "total": total, "supplier": _sup}


def parse_pdf_invoice(raw, name="invoice.pdf"):
    """Розбір PDF-накладної постачальника. Дістаємо таблиці/текст через pdfplumber,
    перетворюємо на сітку рядків і віддаємо в parse_generic_invoice (уся логіка розбору
    шапки/позицій/підсумку/реквізитів переюзається без дублювання).
    Якщо PDF — скан (немає тексту), повернемо порожні lines: чернетка створиться з файлом,
    позиції додаються вручну."""
    import io as _io
    try:
        import pdfplumber
    except Exception as e:  # noqa
        return {"lines": [], "invoice_number": None, "invoice_date": None, "total": None,
                "supplier": {}, "_pdf_error": "pdfplumber недоступний: %s" % e}

    grid = []
    try:
        with pdfplumber.open(_io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                # 1) явні таблиці
                for tbl in (page.extract_tables() or []):
                    for row in tbl:
                        grid.append(["" if c is None else str(c).replace("\n", " ").strip() for c in row])
                # 2) плюс сирий текст рядками (щоб шапка/реквізити/номер теж потрапили)
                txt = page.extract_text() or ""
                for ln in txt.split("\n"):
                    ln = ln.strip()
                    if ln:
                        # ділимо на «колонки» за 2+ пробілами
                        import re as _re
                        cells = _re.split(r"\s{2,}", ln)
                        grid.append([c.strip() for c in cells])
    except Exception as e:  # noqa
        return {"lines": [], "invoice_number": None, "invoice_date": None, "total": None,
                "supplier": {}, "_pdf_error": "не вдалося прочитати PDF: %s" % e}

    if not grid:
        return {"lines": [], "invoice_number": None, "invoice_date": None, "total": None, "supplier": {}}

    # Пакуємо сітку в .xlsx у памʼяті і віддаємо в універсальний парсер
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in grid:
            ws.append([(c[:32000] if isinstance(c, str) else c) for c in row])
        buf = _io.BytesIO()
        wb.save(buf)
        res = parse_generic_invoice(buf.getvalue(), "invoice.xlsx")
    except Exception as e:  # noqa
        return {"lines": [], "invoice_number": None, "invoice_date": None, "total": None,
                "supplier": {}, "_pdf_error": "розбір таблиці PDF не вдався: %s" % e}
    return res
