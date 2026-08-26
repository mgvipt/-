# -*- coding: utf-8 -*-
"""Import paid Instagram follows from the scheduled Meta Ads XLSX email.

Meta's Marketing API does not expose this metric. Meta only supplies it in an
Ads Manager export, therefore the downloader deliberately uses a separate,
persistent browser session that Oleg authorises once. Gmail credentials are
reused from the existing read-only invoices integration; no password or cookie
is stored in the CRM database.
"""
from __future__ import annotations

import email
import imaplib
import io
import os
import re
import time
from datetime import date, datetime
from email.header import decode_header
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


META_SENDER = "advertise-noreply@support.facebook.com"
DAILY_REPORT_MARKER = "wallcov"
DOWNLOAD_DIR = Path(os.environ.get("META_REPORT_DOWNLOAD_DIR", "/reports"))
BROWSER_URL = os.environ.get("META_REPORT_BROWSER_URL", "http://meta-report-browser:4444/wd/hub")


class MetaReportAuthRequired(RuntimeError):
    """The technical browser has not been authorised in Meta Ads Manager yet."""


def _decode(value: str | None) -> str:
    parts = []
    for value_part, encoding in decode_header(value or ""):
        if isinstance(value_part, bytes):
            parts.append(value_part.decode(encoding or "utf-8", "ignore"))
        else:
            parts.append(value_part)
    return "".join(parts)


def _plain_body(message) -> str:
    for part in message.walk():
        if part.get_content_type() == "text/plain" and not part.get_filename():
            payload = part.get_payload(decode=True)
            if payload:
                return payload.decode(part.get_content_charset() or "utf-8", "ignore")
    return ""


def _download_url(body: str) -> str:
    found = re.search(r"https://www\.facebook\.com/ads/report_builder/export/download_report/\?[^\s<>]+", body)
    if not found:
        return ""
    url = found.group(0).rstrip(".,;)")
    parsed = urlparse(url)
    if parsed.netloc != "www.facebook.com" or not parsed.path.startswith("/ads/report_builder/export/download_report/"):
        return ""
    return url


def _normal(value) -> str:
    return re.sub(r"[^a-zа-яіїєґ0-9]+", "", str(value or "").casefold())


def _column(headers, *needles: str) -> int | None:
    for index, header in enumerate(headers):
        normal = _normal(header)
        if all(needle in normal for needle in needles):
            return index
    return None


def _first_column(headers, choices: tuple[tuple[str, ...], ...]) -> int | None:
    for choice in choices:
        found = _column(headers, *choice)
        if found is not None:
            return found
    return None


def _parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    value = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y", "%d/%m/%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    return None


def _parse_number(value) -> int | None:
    if isinstance(value, (int, float)):
        return max(0, int(value))
    text = str(value or "").strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return max(0, int(float(text)))
    except ValueError:
        return None


def parse_report(raw: bytes) -> list[dict]:
    """Return normalised day/campaign/adset/ad follow rows from a Meta XLSX."""
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(values)]
    except StopIteration:
        return []

    day_col = _first_column(headers, (("date",), ("дата",)))
    campaign_col = _first_column(headers, (("campaign", "name"), ("кампан", "наз")))
    adset_col = _first_column(headers, (("adset", "name"), ("adset",), ("group", "name"), ("груп", "объяв")))
    ad_col = next((index for index, header in enumerate(headers)
                   if (_normal(header) == "adname" or "объявлен" in _normal(header))), None)
    follows_col = _first_column(headers, (("instagram", "follow"), ("instagram", "подпис"), ("instagram", "підпис")))
    if day_col is None or follows_col is None:
        raise ValueError("В XLSX Meta не найдены колонки «Дата» и «Instagram Follows»")

    rows = []
    for source in values:
        source = list(source)
        if len(source) <= max(day_col, follows_col):
            continue
        day = _parse_date(source[day_col])
        follows = _parse_number(source[follows_col])
        if day is None or follows is None:
            continue
        item = {
            "date": day,
            "campaign_name": str(source[campaign_col] or "").strip()[:255] if campaign_col is not None and len(source) > campaign_col else "",
            "adset_name": str(source[adset_col] or "").strip()[:255] if adset_col is not None and len(source) > adset_col else "",
            "ad_name": str(source[ad_col] or "").strip()[:255] if ad_col is not None and len(source) > ad_col else "",
            "follows": follows,
        }
        # Meta sometimes writes a visual total row below the actual report.
        if _normal(item["campaign_name"]) in {"total", "итого", "всего"}:
            continue
        rows.append(item)
    return rows


def _download(url: str) -> bytes:
    """Download through the dedicated Selenium Chrome, never via a user cookie."""
    from selenium import webdriver
    from selenium.common.exceptions import WebDriverException

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    before = {path.name: path.stat().st_mtime_ns for path in DOWNLOAD_DIR.glob("*") if path.is_file()}
    options = webdriver.ChromeOptions()
    options.add_experimental_option("prefs", {
        "download.default_directory": str(DOWNLOAD_DIR),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    })
    driver = None
    try:
        driver = webdriver.Remote(command_executor=BROWSER_URL, options=options)
        try:
            driver.execute_cdp_cmd("Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": str(DOWNLOAD_DIR)})
        except WebDriverException:
            pass
        driver.get(url)
        time.sleep(2)
        current_url = (driver.current_url or "").casefold()
        if "/login" in current_url or "/checkpoint" in current_url:
            raise MetaReportAuthRequired("Нужно один раз войти в Meta в техническом браузере CRM")
        deadline = time.monotonic() + 45
        while time.monotonic() < deadline:
            files = [path for path in DOWNLOAD_DIR.glob("*") if path.is_file() and not path.name.endswith(".crdownload")]
            fresh = [path for path in files if before.get(path.name) != path.stat().st_mtime_ns]
            if fresh:
                newest = max(fresh, key=lambda path: path.stat().st_mtime_ns)
                data = newest.read_bytes()
                if data[:2] == b"PK":
                    newest.unlink(missing_ok=True)
                    return data
                raise ValueError("Meta отдала не XLSX-файл")
            time.sleep(1)
        raise ValueError("Meta не отдала XLSX за 45 секунд")
    except WebDriverException as exc:
        raise RuntimeError("Технический браузер Meta недоступен") from exc
    finally:
        if driver is not None:
            driver.quit()


def _settings():
    from .models import IntegrationSettings

    mailbox = IntegrationSettings.objects.filter(provider="email_invoices", is_active=True).first()
    state, _ = IntegrationSettings.objects.get_or_create(
        provider="meta_paid_follows", defaults={"is_active": True, "config": {"done_uids": []}},
    )
    return mailbox, state


def import_reports(*, backfill: bool = False, dry_run: bool = False, limit: int = 100) -> dict:
    """Read new scheduled report emails and import their paid-follower rows.

    Default mode accepts only the daily Wallcov report. ``backfill`` is a
    separate explicit operation for historic reports, so it can be counted and
    approved before a large first import.
    """
    from apps.crm.models import MetaPaidFollowStat

    mailbox, state = _settings()
    cfg = (mailbox.config if mailbox else {}) or {}
    state_cfg = dict(state.config or {})
    user = str(cfg.get("email") or "").strip()
    password = str(cfg.get("app_password") or "").strip()
    host = str(cfg.get("imap_host") or "imap.gmail.com").strip()
    if not user or not password:
        return {"error": "mailbox_not_configured"}

    done = {str(value) for value in state_cfg.get("done_uids") or []}
    client = imaplib.IMAP4_SSL(host)
    client.login(user, password)
    client.select("INBOX", readonly=True)
    try:
        _, data = client.uid("search", None, "FROM", '"%s"' % META_SENDER)
        uids = sorted((int(value) for value in (data[0].split() if data and data[0] else [])))[-limit:]
        scanned = downloaded = imported = rows_seen = 0
        errors = []
        for uid in uids:
            uid_text = str(uid)
            if uid_text in done:
                continue
            _, raw = client.uid("fetch", uid_text, "(RFC822)")
            if not raw or not raw[0]:
                continue
            message = email.message_from_bytes(raw[0][1])
            body = _plain_body(message)
            marker = _normal(body)
            url = _download_url(body)
            if not url or "экспортированныйфайлготовдляскачивания" not in marker:
                continue
            if not backfill and DAILY_REPORT_MARKER not in marker:
                continue
            scanned += 1
            try:
                rows = parse_report(_download(url))
                downloaded += 1
            except MetaReportAuthRequired as exc:
                state_cfg["last_error"] = str(exc)
                errors.append(str(exc))
                break
            except Exception as exc:  # keep this mail pending for a safe retry
                state_cfg["last_error"] = str(exc)[:500]
                errors.append(str(exc)[:200])
                continue
            if not rows:
                done.add(uid_text)
                continue
            rows_seen += len(rows)
            if not dry_run:
                for row in rows:
                    MetaPaidFollowStat.objects.update_or_create(
                        date=row["date"], campaign_name=row["campaign_name"],
                        adset_name=row["adset_name"], ad_name=row["ad_name"],
                        defaults={"follows": row["follows"], "report_uid": uid_text},
                    )
                    imported += 1
                done.add(uid_text)
        state_cfg["done_uids"] = sorted(done, key=int)[-200:]
        state_cfg["last_checked_at"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
        if not errors:
            state_cfg.pop("last_error", None)
        state.config = state_cfg
        state.is_active = True
        state.save(update_fields=["config", "is_active", "updated_at"])
        return {"scanned": scanned, "downloaded": downloaded, "rows": rows_seen, "imported": imported,
                "dry_run": dry_run, "errors": errors}
    finally:
        client.logout()


# ═══════════════════════════════════════════════════════════════════════════
# Джерело №2 (Claude, 26.08.2026): анонімне share-посилання Ads Reporting.
#
# ЧОМУ (для Codex і майбутніх сесій): розсилку звітів ПОШТОЮ Meta прибрала з
# нового кабінету Ads Reporting — перевірено вручну 26.08.2026, в UI лишився
# тільки разовий експорт файлом. Тож листи від advertise-noreply самі не
# приходитимуть, а download_report-посилання з них вимагає залогіненого
# браузера (звідси Selenium вище). НАТОМІСТЬ «Поділитися → доступ за
# посиланням» віддає сторінку зі вбудованими даними ПОВНІСТЮ АНОНІМНО —
# достатньо браузерних заголовків, без cookie/логіна/Selenium.
#
# Обмеження: у HTML вбудовані перші ~50 рядків (сортування за датою desc) =
# останні 3-4 дні. Для щоденного крону цього достатньо (upsert перекриває
# пізню атрибуцію). Історію разово вантажимо CSV-файлом (--csv-file).
# Посилання діє 30 днів (expire пишемо у state і попереджаємо заздалегідь).
# Звіт у кабінеті: «CRM IG Follows Daily» (акаунт Wallcov, розбивка
# день+кампанія+campaign_id, метрика «Підписки в Instagram»).
# ═══════════════════════════════════════════════════════════════════════════

SHARE_URL_ENV = "META_ADS_SHARE_REPORT_URL"
_SHARE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ru-RU,ru;q=0.9,uk;q=0.8,en-US;q=0.7",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
}


def share_report_url() -> str:
    return os.environ.get(SHARE_URL_ENV, "").strip()


def fetch_share_html(url: str | None = None, timeout: int = 60) -> str:
    import requests

    target = (url or share_report_url())
    if not target:
        raise ValueError("META_ADS_SHARE_REPORT_URL is not configured")
    session = requests.Session()
    response = session.get(target, headers=_SHARE_HEADERS, timeout=timeout, allow_redirects=True)
    response.raise_for_status()
    if "adsviewreport" not in (response.url or "") and "dimensions" not in response.text:
        raise ValueError("Share link did not return the report page (login wall or expired link?)")
    return response.text


def _iter_row_objects(html: str):
    """Yield each embedded {"dimensions":[...],"metrics":[...]} object as parsed JSON."""
    import json as _json

    for match in re.finditer(r'\{"dimensions":\[\{"value":"(\d{4}-\d\d-\d\d)"', html):
        start = match.start()
        depth = 0
        end = None
        for index in range(start, min(start + 20000, len(html))):
            char = html[index]
            if char == "{":
                depth += 1
            elif char == "}":
                depth -= 1
                if depth == 0:
                    end = index + 1
                    break
        if end is None:
            continue
        try:
            yield _json.loads(html[start:end])
        except ValueError:
            continue


def parse_share_html(html: str) -> tuple[list[dict], dict]:
    """Return campaign-level follow rows + report metadata from the share page.

    Індекс метрики НЕ хардкодиться: береться зі схеми полів у тому ж HTML
    (список [{"name":"delivery_info"...}]) — переживе зміну складу колонок.
    """
    schema_match = re.search(r'\[\{"name":"delivery_info".*?\]', html)
    if not schema_match:
        raise ValueError("Share page: field schema not found (layout changed?)")
    field_names = re.findall(r'\{"name":"([a-z_0-9:.]+)"', schema_match.group(0))
    try:
        follows_index = field_names.index("instagram_profile_follow_v2")
    except ValueError:
        raise ValueError("Share page: instagram_profile_follow_v2 missing from schema")

    meta: dict = {}
    meta_match = re.search(r'"reportName":"([^"]*)","updateTime":(\d+),"expireTime":(\d+)', html)
    if meta_match:
        meta = {
            "report_name": meta_match.group(1),
            "update_time": int(meta_match.group(2)),
            "expire_time": int(meta_match.group(3)),
        }

    rows: list[dict] = []
    for row in _iter_row_objects(html):
        dims = [d.get("value") for d in row.get("dimensions") or []]
        if len(dims) < 3:
            continue
        day, campaign_name, campaign_id = dims[0], dims[1], dims[2]
        # рядок кампанії без дубля: беремо той, де третя розмірність = campaign_id
        if campaign_name == "__summary__" or not str(campaign_id or "").isdigit():
            continue
        metrics = row.get("metrics") or []
        if follows_index >= len(metrics):
            continue
        value = _parse_number((metrics[follows_index] or {}).get("value"))
        parsed_day = _parse_date(day)
        if parsed_day is None or value is None:
            continue
        rows.append({
            "date": parsed_day,
            "campaign_name": str(campaign_name or "").strip()[:255],
            "adset_name": "",
            "ad_name": "",
            "follows": value,
        })
    return rows, meta


def import_share_report(*, dry_run: bool = False) -> dict:
    """Fetch the anonymous share page and upsert campaign-level follow rows."""
    from apps.crm.models import MetaPaidFollowStat
    from .models import IntegrationSettings

    state, _ = IntegrationSettings.objects.get_or_create(
        provider="meta_paid_follows", defaults={"is_active": True, "config": {}},
    )
    state_cfg = dict(state.config or {})
    try:
        html = fetch_share_html()
        rows, meta = parse_share_html(html)
    except Exception as exc:
        state_cfg["share_last_error"] = str(exc)[:500]
        state.config = state_cfg
        state.save(update_fields=["config", "updated_at"])
        return {"source": "share", "error": str(exc)[:200]}

    imported = 0
    if not dry_run:
        for row in rows:
            MetaPaidFollowStat.objects.update_or_create(
                date=row["date"], campaign_name=row["campaign_name"],
                adset_name="", ad_name="",
                defaults={"follows": row["follows"], "report_uid": "share"},
            )
            imported += 1

    warning = ""
    expire_time = meta.get("expire_time")
    if expire_time:
        days_left = (datetime.utcfromtimestamp(expire_time) - datetime.utcnow()).days
        state_cfg["share_expires_at"] = datetime.utcfromtimestamp(expire_time).isoformat(timespec="seconds") + "Z"
        if days_left <= 7:
            warning = ("Share-посилання Ads-звіту спливає через %d дн. — Олег має перевидати його: "
                       "Ads Reporting → CRM IG Follows Daily → Поділитися → строк дії" % days_left)
            state_cfg["share_expire_warning"] = warning
        else:
            state_cfg.pop("share_expire_warning", None)
    state_cfg["share_last_sync"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    state_cfg.pop("share_last_error", None)
    state.config = state_cfg
    state.is_active = True
    state.save(update_fields=["config", "is_active", "updated_at"])
    dates = sorted({row["date"].isoformat() for row in rows})
    return {"source": "share", "rows": len(rows), "imported": imported, "dates": dates,
            "dry_run": dry_run, **({"warning": warning} if warning else {})}


def import_csv_file(path: str, *, dry_run: bool = False) -> dict:
    """One-off history backfill from an Ads Reporting CSV export («CRM IG Follows Daily»)."""
    import csv

    from apps.crm.models import MetaPaidFollowStat

    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        try:
            headers = next(reader)
        except StopIteration:
            return {"source": "csv", "error": "empty file"}
        day_col = _first_column(headers, (("день",), ("date",), ("дата",)))
        campaign_col = _first_column(headers, (("кампан", "наз"), ("campaign", "name")))
        follows_col = _first_column(headers, (("instagram", "подпис"), ("instagram", "підпис"), ("instagram", "follow")))
        if day_col is None or follows_col is None:
            return {"source": "csv", "error": "columns not found: %s" % headers}
        rows_seen = imported = 0
        for source in reader:
            if len(source) <= max(day_col, follows_col):
                continue
            day = _parse_date(source[day_col])
            follows = _parse_number(source[follows_col])
            campaign = str(source[campaign_col] or "").strip()[:255] if campaign_col is not None and len(source) > campaign_col else ""
            if day is None or follows is None or not campaign or campaign.casefold() in {"все", "всего", "итого", "total"}:
                continue
            rows_seen += 1
            if not dry_run:
                MetaPaidFollowStat.objects.update_or_create(
                    date=day, campaign_name=campaign, adset_name="", ad_name="",
                    defaults={"follows": follows, "report_uid": "csv-backfill"},
                )
                imported += 1
    return {"source": "csv", "rows": rows_seen, "imported": imported, "dry_run": dry_run}
