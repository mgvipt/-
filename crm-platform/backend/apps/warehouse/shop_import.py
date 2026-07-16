from __future__ import annotations

import hashlib
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook

from .models import Product, ShopCatalogImportBatch, ShopCatalogImportRow
from .shop_sync import infer_sample_metadata, queue_product_sync


MATERIAL_SHEET = "Материалы"
SAMPLE_SHEET = "Тест-наборы"
SHEETS = (MATERIAL_SHEET, SAMPLE_SHEET)


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def as_decimal(value):
    value = clean(value)
    if value is None:
        return None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def as_bool(value):
    return str(value or "").strip().lower() in {"да", "так", "yes", "true", "1", "+"}


def split_list(value):
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in text.replace(";", ",").split(",") if part.strip()]


def external_id(value):
    value = clean(value)
    if value is None:
        return ""
    return str(value).removesuffix(".0")


def material_specs(columns):
    return {
        "source_sheet": MATERIAL_SHEET,
        "source_columns": columns,
        "price_model": clean(columns.get("Цена: модель")),
        "price_per_kg": clean(columns.get("Цена за кг, грн")),
        "price_per_m2": clean(columns.get("Цена за м², грн")),
        "price_from": clean(columns.get("Цена ОТ, грн")),
        "consumption_min": clean(columns.get("Расход min, кг/м²")),
        "consumption_max": clean(columns.get("Расход max, кг/м²")),
        "price_per_m2_min": clean(columns.get("Расчётная цена min, грн/м²")),
        "price_per_m2_max": clean(columns.get("Расчётная цена max, грн/м²")),
        "finish": clean(columns.get("Финиш")),
        "washable": clean(columns.get("Можно мыть?")),
        "durability": clean(columns.get("Износостойкость")),
        "packaging": clean(columns.get("Фасовка / отгрузка")),
    }


def build_proposal(sheet_name, columns, product=None):
    source_id = external_id(columns.get("ID / SKU") if sheet_name == MATERIAL_SHEET else columns.get("ID материала"))
    if sheet_name == MATERIAL_SHEET:
        client_name = clean(columns.get("Клиентское название")) or clean(columns.get("Внутреннее название")) or (product.name if product else "")
        specs = material_specs(columns)
        benefits = [str(value) for value in (
            columns.get("Финиш"), columns.get("Можно мыть?"), columns.get("Износостойкость")
        ) if clean(value)]
        return {
            "price": clean(columns.get("Цена за кг, грн")),
            "shop_parent_name": client_name,
            "shop_slug": (product.shop_slug if product and product.shop_slug else slugify(client_name, allow_unicode=False) or f"material-{source_id}"),
            "shop_group_key": (product.shop_group_key if product and product.shop_group_key else f"material-{source_id}"),
            "shop_short_description": clean(columns.get("Короткое описание для клиента")) or "",
            "shop_full_description": clean(columns.get("Когда рекомендовать")) or clean(columns.get("Для какого запроса подходит")) or "",
            "shop_effect": clean(columns.get("Эффект / ассоциация")) or "",
            "shop_rooms": split_list(columns.get("Где использовать")),
            "shop_beginner": as_bool(columns.get("Можно наносить самому?")),
            "shop_benefits": benefits,
            "shop_variant_type": "product",
            "shop_variant_order": 1,
            "shop_variant_name": "Основной товар",
            "shop_specs": specs,
        }

    inferred = infer_sample_metadata(product.name if product else str(columns.get("Внутреннее название набора") or ""))
    client_name = clean(columns.get("Клиентское название набора")) or clean(columns.get("Внутреннее название набора")) or (product.name if product else "")
    return {
        "price": clean(columns.get("Цена, грн")),
        "shop_parent_name": product.shop_parent_name if product and product.shop_parent_name else inferred["shop_parent_name"],
        "shop_slug": product.shop_slug if product and product.shop_slug else inferred["shop_slug"],
        "shop_group_key": product.shop_group_key if product and product.shop_group_key else inferred["shop_group_key"],
        "shop_short_description": clean(columns.get("Когда предлагать")) or "",
        "shop_full_description": clean(columns.get("Фраза продажи для агента")) or "",
        "shop_contents": clean(columns.get("Что входит")) or "",
        "shop_variant_type": "sample",
        "shop_variant_order": product.shop_variant_order if product and product.shop_variant_order else inferred["shop_variant_order"],
        "shop_variant_name": client_name,
        "shop_has_board": product.shop_has_board if product and product.shop_has_board is not None else inferred["shop_has_board"],
        "shop_is_tinted": product.shop_is_tinted if product and product.shop_is_tinted is not None else inferred["shop_is_tinted"],
        "shop_specs": {
            "source_sheet": SAMPLE_SHEET,
            "source_columns": columns,
            "sample_size": clean(columns.get("Размер дощечки / образца")),
            "tinting": clean(columns.get("Тонировка")),
            "public_price_allowed": clean(columns.get("Можно ли озвучивать цену?")),
        },
    }


def validate_row(sheet_name, columns, product=None):
    warnings, errors = [], []
    if product is None:
        errors.append("Товар с таким ID не найден в номенклатуре CRM")
    price_header = "Цена за кг, грн" if sheet_name == MATERIAL_SHEET else "Цена, грн"
    source_price = as_decimal(columns.get(price_header))
    if source_price is None or source_price <= 0:
        errors.append("В таблице не указана корректная цена")
    elif product is not None and as_decimal(product.price) != source_price:
        warnings.append(f"Цена отличается: CRM {product.price} грн → таблица {source_price} грн")

    if sheet_name == MATERIAL_SHEET:
        consumption_min = as_decimal(columns.get("Расход min, кг/м²"))
        consumption_max = as_decimal(columns.get("Расход max, кг/м²"))
        price_min = as_decimal(columns.get("Расчётная цена min, грн/м²"))
        price_max = as_decimal(columns.get("Расчётная цена max, грн/м²"))
        if consumption_min is not None and consumption_max is not None and consumption_min > consumption_max:
            errors.append("Минимальный расход больше максимального — проверьте значения")
        if all(value is not None for value in (source_price, consumption_min, consumption_max, price_min, price_max)):
            expected_min = (source_price * consumption_min).quantize(Decimal("0.01"))
            expected_max = (source_price * consumption_max).quantize(Decimal("0.01"))
            if abs(expected_min - price_min) > Decimal("0.02") or abs(expected_max - price_max) > Decimal("0.02"):
                errors.append(f"Расчёт цены за м² не сходится: должно быть {expected_min}–{expected_max} грн")
    return warnings, errors


def serialize_row(row):
    product = row.product
    return {
        "id": row.id,
        "sheet_name": row.sheet_name,
        "source_row": row.source_row,
        "external_id": row.external_id,
        "product_id": row.product_id,
        "product_name": product.name if product else "",
        "sku": product.sku if product else "",
        "current_price": str(product.price) if product else None,
        "source_data": row.source_data,
        "proposed_data": row.proposed_data,
        "warnings": row.warnings,
        "errors": row.errors,
        "applied_at": row.applied_at.isoformat() if row.applied_at else None,
    }


def batch_summary(rows):
    rows = list(rows)
    return {
        "rows": len(rows),
        "matched": sum(bool(row.product_id) for row in rows),
        "unmatched": sum(not row.product_id for row in rows),
        "warnings": sum(bool(row.warnings) for row in rows),
        "errors": sum(bool(row.errors) for row in rows),
        "price_changes": sum(any("Цена отличается" in message for message in row.warnings) for row in rows),
        "applied": sum(bool(row.applied_at) for row in rows),
    }


@transaction.atomic
def import_workbook(uploaded_file, user=None):
    raw = uploaded_file.read()
    workbook = load_workbook(BytesIO(raw), data_only=True)
    missing_sheets = [name for name in SHEETS if name not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError("В файле нет листов: " + ", ".join(missing_sheets))
    batch = ShopCatalogImportBatch.objects.create(
        source_name=(uploaded_file.name or "catalog.xlsx")[:255],
        source_hash=hashlib.sha256(raw).hexdigest(), created_by=user,
    )
    created = []
    for sheet_name in SHEETS:
        sheet = workbook[sheet_name]
        headers = [clean(cell.value) for cell in sheet[1]]
        id_header = "ID / SKU" if sheet_name == MATERIAL_SHEET else "ID материала"
        for source_row in range(2, sheet.max_row + 1):
            columns = {headers[index]: clean(sheet.cell(source_row, index + 1).value) for index in range(len(headers)) if headers[index]}
            source_id = external_id(columns.get(id_header))
            if not source_id:
                continue
            product = Product.objects.filter(bitrix_id=source_id).first()
            warnings, errors = validate_row(sheet_name, columns, product)
            created.append(ShopCatalogImportRow.objects.create(
                batch=batch, sheet_name=sheet_name, source_row=source_row,
                external_id=source_id, product=product, source_data=columns,
                proposed_data=build_proposal(sheet_name, columns, product),
                warnings=warnings, errors=errors,
            ))
    batch.summary = batch_summary(created)
    batch.save(update_fields=["summary", "updated_at"])
    return batch


@transaction.atomic
def update_import_row(row, columns):
    product = row.product
    warnings, errors = validate_row(row.sheet_name, columns, product)
    row.source_data = columns
    row.proposed_data = build_proposal(row.sheet_name, columns, product)
    row.warnings = warnings
    row.errors = errors
    row.save(update_fields=["source_data", "proposed_data", "warnings", "errors", "updated_at"])
    row.batch.summary = batch_summary(row.batch.rows.select_related("product"))
    row.batch.save(update_fields=["summary", "updated_at"])
    return row


@transaction.atomic
def apply_rows(batch, row_ids):
    # Product nullable: PostgreSQL не дозволяє FOR UPDATE на nullable-боці OUTER JOIN.
    rows = list(batch.rows.select_for_update().filter(id__in=row_ids))
    blocked = [row.id for row in rows if row.errors or not row.product_id]
    if blocked:
        raise ValueError("Сначала исправьте строки: " + ", ".join(map(str, blocked)))
    for row in rows:
        product = row.product
        proposal = row.proposed_data
        for field in (
            "shop_parent_name", "shop_slug", "shop_group_key", "shop_short_description", "shop_full_description",
            "shop_effect", "shop_rooms", "shop_beginner", "shop_benefits", "shop_variant_type", "shop_variant_order",
            "shop_variant_name", "shop_has_board", "shop_is_tinted", "shop_contents", "shop_specs",
        ):
            if field in proposal:
                setattr(product, field, proposal[field])
        product.price = as_decimal(proposal.get("price")) or product.price
        product.shop_managed = True
        product.save()
        if product.shop_enabled:
            queue_product_sync(product)
        row.applied_at = timezone.now()
        row.save(update_fields=["applied_at", "updated_at"])
    total = batch.rows.count()
    applied = batch.rows.filter(applied_at__isnull=False).count()
    batch.status = "applied" if total and applied == total else "partially_applied"
    batch.summary = batch_summary(batch.rows.select_related("product"))
    batch.save(update_fields=["status", "summary", "updated_at"])
    return rows
